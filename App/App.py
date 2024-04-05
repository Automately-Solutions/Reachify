from rich import print, box
from rich.panel import Panel
import pandas as pd
import requests
from bs4 import BeautifulSoup
import re
from instagrapi import Client
from instagrapi.exceptions import UserNotFound
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

logger = logging.getLogger()

def login_user(username, password):
    """
    Login to Instagram with the provided username and password.
    """
    cl = Client()
    if cl.login(username, password):
        return cl
    else:
        raise Exception("Login failed")

def scrape_social_media_links(website_url):
    facebook_links = []
    gmail_addresses = []
    try:
        response = requests.get(website_url, timeout=10)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            for link in soup.find_all('a', href=True):
                if "facebook.com" in link['href']:
                    facebook_links.append(link['href'])
            for email in re.findall(r"[a-zA-Z0-9_.+-]+@gmail.com", soup.text):
                gmail_addresses.append(email)
    except requests.RequestException as e:
        print(f"Error fetching {website_url}: {e}")
    return facebook_links, gmail_addresses

def send_instagram_messages(websites, cl):
    messages_sent = 0
    for i, url in enumerate(websites, start=2):  # start=2 assumes headers in row 1
        facebook_links, gmail_addresses = scrape_social_media_links(url)
        print(Panel(f"Facebook Links: {facebook_links}", title="Facebook Links", border_style="bold blue", box=box.SQUARE))
        print(Panel(f"Gmail Addresses: {gmail_addresses}", title="Gmail Addresses", border_style="bold yellow", box=box.SQUARE))
        
        # Now, onto Instagram messaging
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            for link in soup.find_all('a', href=True):
                if "instagram.com" in link['href']:
                    username = extract_instagram_username(link['href'])
                    if username:
                        try:
                            user_id = cl.user_id_from_username(username)
                            message = f"Hey {username},\n\nImpressed by the range of services, especially as summer heats up the demand. At WordSmith, we offer expert digital marketing with a twist: no payment until you see results. Ready to make this summer your most profitable one? Let's chat."                            
                            cl.direct_send(message, [user_id])
                            ws['A' + str(i)].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # A is the column with websites, adjust accordingly
                            print(Panel.fit(f"Message sent to {username}", border_style="bold green", box=box.SQUARE))
                            messages_sent += 1
                            break
                        except UserNotFound:
                            print(Panel.fit(f"Instagram user {username} not found.", title="Error", title_slign="left", border_style="bold red", box=box.SQUARE))
        wb.save(file_path)
    print(Panel.fit(f"Successfully sent messages: {messages_sent}", border_style="bold green", box=box.SQUARE))

def extract_instagram_username(instagram_url):
    match = re.search(r"instagram.com/([^/?#&]+)", instagram_url)
    return match.group(1) if match else None

# Load the .xlsx file
file_path = 'Client Spreadsheet.xlsx'
wb = load_workbook(filename=file_path)
ws = wb.active

# Assuming the website links are in column 'F'
websites = [cell.value for cell in ws['F'][1:]]  # Skip header

# Instagram client setup
username = "username"
password = "password"
cl = login_user(username, password)

send_instagram_messages(websites, cl)
